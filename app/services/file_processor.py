import io
import os
from datetime import datetime
from typing import Optional

import pandas as pd
from scipy import stats
from difflib import SequenceMatcher
from fastapi import UploadFile

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader

from app.schemas.summarize import FileProcessResult
from app.utils.validation import validate_dataframe_size
from app.services.ai_summarizer import generate_summary,generate_eda_insight

TEMP_DIR = "temp"

# pdf file processing function
async def process_pdf_file(file: UploadFile) -> FileProcessResult:
    content = await file.read()
    reader = PdfReader(io.BytesIO(content))
    raw_text = ""
    for page in reader.pages:
        raw_text += page.extract_text() or ""

    if not raw_text.strip():
        return FileProcessResult(
            summary="Tidak ada teks yang dapat diekstrak dari file PDF.",
            output_bytes=b"",
            output_filename=f"summarized_{file.filename.replace(' ', '_')}.pdf"
        )

    summary = await generate_summary(raw_text)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    c.setFont("Helvetica", 12)
    c.drawString(50, 800, "Ringkasan Otomatis:")

    y = 780
    for line in summary.splitlines():
        c.drawString(50, y, line)
        y -= 18
        if y < 50:
            c.showPage()
            y = 800

    c.save()
    output_bytes = buffer.getvalue()

    return FileProcessResult(
        summary=summary,
        output_bytes=output_bytes,
        output_filename=f"summarized_{file.filename.replace(' ', '_')}.pdf"
    )



async def process_excel_file(
    file: UploadFile,
    filter_field: Optional[str] = None,
    filter_value: Optional[str] = None
) -> FileProcessResult:
    filename = file.filename.lower()
    content = await file.read()
    if filename.endswith(".csv"):
        try:
            df = pd.read_csv(io.BytesIO(content), encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(io.BytesIO(content), encoding="latin-1")
    else:
        df = pd.read_excel(io.BytesIO(content))

    # 
    validate_dataframe_size(df, max_rows=10000, max_cols=15)

    original_rows = df.shape[0]
    if filter_field and filter_value and filter_field in df.columns:
        df = df[df[filter_field].astype(str).str.lower() == filter_value.lower()]
    filtered_rows = df.shape[0]

    df.columns = [col.strip().replace(" ", "_").lower() for col in df.columns]
    df.dropna(how='all', axis=1, inplace=True)
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].astype(str).str.strip().str.lower()

    for col in df.select_dtypes(include='object').columns:
        unique_vals = df[col].dropna().unique().tolist()
        merged = {}
        for i in range(len(unique_vals)):
            for j in range(i + 1, len(unique_vals)):
                a, b = unique_vals[i], unique_vals[j]
                if a != b and SequenceMatcher(None, a, b).ratio() > 0.95:
                    merged[b] = a
        df[col] = df[col].replace(merged)

    feature_types = []
    for col in df.columns:
        nunique = df[col].nunique()
        dtype = df[col].dtype
        if 'date' in col or 'time' in col:
            ftype = 'datetime'
        elif nunique == original_rows:
            ftype = 'id'
        elif dtype in ['float64', 'int64'] and nunique > 10:
            ftype = 'numerical'
        elif dtype in ['object', 'category'] or nunique <= 10:
            ftype = 'categorical'
        else:
            ftype = 'unknown'
        feature_types.append((col, ftype, nunique, str(dtype)))

    missing_summary = df.isnull().sum()
    missing_pct = (missing_summary / len(df) * 100).round(2)
    fillna_strategy = {}
    for col in df.columns:
        if df[col].dtype in ['float64', 'int64']:
            if missing_pct[col] < 5:
                df[col].fillna(df[col].median(), inplace=True)
                fillna_strategy[col] = "median"
            elif missing_pct[col] < 20:
                df[col].fillna(df[col].mean(), inplace=True)
                fillna_strategy[col] = "mean"
            else:
                fillna_strategy[col] = "high missing rate - manual review"
        elif df[col].dtype == 'object':
            df[col].fillna("unknown", inplace=True)
            fillna_strategy[col] = "fill with 'unknown'"

    numeric_summary = df.describe().T[['mean', 'std', 'min', 'max']].round(2)
    correlation_matrix = df.corr(numeric_only=True).round(2)
    z_scores = stats.zscore(df.select_dtypes(include='number'), nan_policy='omit')
    outliers = (abs(z_scores) > 3)
    df_outliers = df[outliers.any(axis=1)]

    quality_flags = []
    for col in df.columns:
        score = "OK"
        if missing_pct[col] > 30:
            score = "❌ High Missing"
        elif df[col].duplicated().sum() > 0:
            score = "⚠️ Duplicates Exist"
        elif df[col].nunique() == 1:
            score = "⚠️ Constant Value"
        quality_flags.append((col, missing_summary[col], missing_pct[col], score))

    insight_text = await generate_eda_insight(df.head(100).to_csv(index=False))

    filename_out = f"transformed_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename.replace(' ', '_')}"
    output_path = os.path.join(TEMP_DIR, filename_out)
    os.makedirs(TEMP_DIR, exist_ok=True)

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Cleaned Data"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)
    table_range = f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}"
    table = Table(displayName="CleanedTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws_data.add_table(table)

    ws_summary = wb.create_sheet("Summary Stats")
    ws_summary.append(["Original Rows", original_rows])
    ws_summary.append(["Filtered Rows", filtered_rows])
    ws_summary.append([])
    for row in dataframe_to_rows(numeric_summary.reset_index(), index=False, header=True):
        ws_summary.append(row)

    ws_corr = wb.create_sheet("Correlation")
    for row in dataframe_to_rows(correlation_matrix.reset_index(), index=False, header=True):
        ws_corr.append(row)

    ws_outliers = wb.create_sheet("Outliers")
    for row in dataframe_to_rows(df_outliers, index=False, header=True):
        ws_outliers.append(row)

    ws_fillna = wb.create_sheet("FillNA Strategy")
    ws_fillna.append(["Column", "Strategy"])
    for col, strategy in fillna_strategy.items():
        ws_fillna.append([col, strategy])

    ws_class = wb.create_sheet("Feature Classification")
    ws_class.append(["Column", "Type", "Unique Values", "Dtype"])
    for row in feature_types:
        ws_class.append(row)

    ws_quality = wb.create_sheet("Data Quality")
    ws_quality.append(["Column", "Missing Count", "Missing %", "Flag"])
    for row in quality_flags:
        ws_quality.append(row)

    vis_sheet = wb.create_sheet("Visualizations")
    row_cursor = 1

    for cat_col in df.select_dtypes(include='object').columns:
        if df[cat_col].nunique() <= 10:
            for num_col in df.select_dtypes(include='number').columns:
                grouped = df.groupby(cat_col)[num_col].agg('mean').reset_index()
                vis_sheet.cell(row=row_cursor, column=1, value=f"BarChart: {num_col} vs {cat_col}")
                for row in dataframe_to_rows(grouped, index=False, header=True):
                    for j, val in enumerate(row):
                        vis_sheet.cell(row=row_cursor + 1, column=j + 1, value=val)
                    row_cursor += 1
                row_cursor += 2

                chart = BarChart()
                chart.title = f"BarChart: {num_col} vs {cat_col}"
                chart.y_axis.title = num_col
                chart.x_axis.title = cat_col
                data = Reference(vis_sheet, min_col=2, min_row=row_cursor - len(grouped) - 1, max_row=row_cursor - 2)
                categories = Reference(vis_sheet, min_col=1, min_row=row_cursor - len(grouped) - 1 + 1, max_row=row_cursor - 2)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                vis_sheet.add_chart(chart, f"E{row_cursor}")
                row_cursor += 15
                break

    for col in df.select_dtypes(include='object').columns:
        if df[col].nunique() <= 6:
            pie_data = df[col].value_counts().reset_index()
            pie_data.columns = [col, 'count']
            vis_sheet.cell(row=row_cursor, column=1, value=f"PieChart: {col}")
            for row in dataframe_to_rows(pie_data, index=False, header=True):
                for j, val in enumerate(row):
                    vis_sheet.cell(row=row_cursor + 1, column=j + 1, value=val)
                row_cursor += 1
            row_cursor += 2
            pie_chart = PieChart()
            pie_chart.title = f"PieChart: {col}"
            data = Reference(vis_sheet, min_col=2, min_row=row_cursor - len(pie_data) - 1, max_row=row_cursor - 2)
            labels = Reference(vis_sheet, min_col=1, min_row=row_cursor - len(pie_data) - 1 + 1, max_row=row_cursor - 2)
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            vis_sheet.add_chart(pie_chart, f"F{row_cursor}")
            row_cursor += 15
            break

    for col in df.columns:
        if 'date' in col or 'time' in col:
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                if df[col].notnull().sum() > 0:
                    df_ts = df[[col]].copy()
                    df_ts['count'] = 1
                    df_ts = df_ts.dropna().groupby(df_ts[col].dt.to_period('M')).count()
                    df_ts.index = df_ts.index.astype(str)
                    ts_data = df_ts.reset_index()
                    ts_data.columns = ['month', 'count']
                    vis_sheet.cell(row=row_cursor, column=1, value=f"LineChart: trend {col}")
                    for row in dataframe_to_rows(ts_data, index=False, header=True):
                        for j, val in enumerate(row):
                            vis_sheet.cell(row=row_cursor + 1, column=j + 1, value=val)
                        row_cursor += 1
                    row_cursor += 2
                    chart = LineChart()
                    chart.title = f"Trend {col}"
                    chart.y_axis.title = "Count"
                    chart.x_axis.title = "Month"
                    data = Reference(vis_sheet, min_col=2, min_row=row_cursor - len(ts_data) - 1, max_row=row_cursor - 2)
                    labels = Reference(vis_sheet, min_col=1, min_row=row_cursor - len(ts_data) - 1 + 1, max_row=row_cursor - 2)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    vis_sheet.add_chart(chart, f"F{row_cursor}")
                    row_cursor += 15
                    break
            except Exception:
                continue

    ws_ai = wb.create_sheet("AI Insight")
    for i, line in enumerate(insight_text.splitlines(), 1):
        ws_ai.cell(row=i, column=1, value=line)

    wb.save(output_path)
    return FileProcessResult(
        summary=f"Excel processed with advanced EDA. Rows before: {original_rows}, after: {filtered_rows}",
        output_filename=filename_out,
        output_path=output_path,
        output_bytes=open(output_path, "rb").read()
    )
